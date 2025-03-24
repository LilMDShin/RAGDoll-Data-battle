import re
import os
import requests
from urllib.parse import urlsplit
from openpyxl import load_workbook
import urllib3

# Disable SSL warnings for testing purposes (not recommended for production)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Path to your .xlsx file
EXCEL_FILE = "Index Legal Official Publications and Certifications data package.xlsx"
SAVE_FOLDER = "Data"  # Folder to save the downloaded files


def sanitize_filename(filename: str) -> str:
    """
    Replace characters that are invalid on typical filesystems with underscores.
    """
    return re.sub(r'[\\/*?:"<>|]', '_', filename)


def extract_filename(response, url):
    """
    Extract the filename from the Content-Disposition header if available.
    Otherwise, fallback to the basename from the URL.
    """
    cd = response.headers.get('content-disposition')
    if cd:
        fname_match = re.findall('filename="?([^"]+)"?', cd)
        if fname_match:
            return fname_match[0]
    return os.path.basename(urlsplit(url).path)


def main():
    os.makedirs(SAVE_FOLDER, exist_ok=True)

    wb = load_workbook(EXCEL_FILE, data_only=False)
    ws = wb.active

    # Process rows assuming headers are in row 1: Name, Type, Year, Format, Link
    for row in ws.iter_rows(min_row=2):
        # Only consider the first 5 columns
        name_cell, type_cell, year_cell, format_cell, link_cell = row[:5]
        row_num = name_cell.row  # For logging purposes

        name = str(name_cell.value) if name_cell.value else ""
        doc_type = str(type_cell.value) if type_cell.value else ""
        year = str(year_cell.value) if year_cell.value else ""
        file_format = str(format_cell.value).lower() if format_cell.value else ""

        # Get the actual hyperlink from the cell if available
        if link_cell.hyperlink:
            link = link_cell.hyperlink.target
        else:
            link = str(link_cell.value)

        if link and link.startswith("http"):
            # Verify that format is either "pdf" or "html"
            if file_format not in ["pdf", "html"]:
                print(f"Skipping row {row_num} because format '{file_format}' is not supported.")
                continue
            # For HTML, we rely on the Excel column "Format" and do not check the link extension

            # Use a HEAD request to try to get the original filename
            try:
                head_response = requests.head(link, allow_redirects=True, verify=False)
                head_response.raise_for_status()
                filename = extract_filename(head_response, link)
            except requests.exceptions.RequestException:
                filename = os.path.basename(urlsplit(link).path)

            filename = sanitize_filename(filename)
            filepath = os.path.join(SAVE_FOLDER, filename)

            # Skip if the file already exists
            if os.path.exists(filepath):
                print(f"Skipping row {row_num} because file already downloaded: {filepath}\n")
                continue

            print(f"Downloading: {link}")
            try:
                response = requests.get(link, stream=True, verify=False)
                response.raise_for_status()

                # Optionally re-extract the filename from the GET response
                filename_get = extract_filename(response, link)
                filename_get = sanitize_filename(filename_get)
                filepath = os.path.join(SAVE_FOLDER, filename_get)
                if os.path.exists(filepath):
                    print(f"Skipping row {row_num} because file already downloaded: {filepath}\n")
                    continue

                with open(filepath, "wb") as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        if chunk:
                            f.write(chunk)
                print(f"Saved to: {filepath}\n")
            except requests.exceptions.RequestException as e:
                print(f"Failed to download row {row_num} with link {link}:")
                print(f"  Name: {name}, Type: {doc_type}, Year: {year}, Format: {file_format}")
                print(f"  Error: {e}\n")
        else:
            print(f"Skipping row {row_num}: no valid URL found: {link}\n")


if __name__ == "__main__":
    main()